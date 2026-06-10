"""Extract settlement (remittance-advice) data from an uploaded file.

Accepts PDF, Excel (.xlsx), or CSV. The file is converted to text, then one
OpenAI structured-output call returns the header fields + per-claim line items.
On any failure we return an empty editable shape so the admin can still fill it
in by hand — mirrors mou_extraction_service.
"""
import csv
import io
import json
import logging

from openai import OpenAI
from PyPDF2 import PdfReader

from app.core.config import settings

logger = logging.getLogger(__name__)


_SCHEMA = {
    "name": "settlement_extract",
    "strict": True,
    "schema": {
        "type": "object",
        "additionalProperties": False,
        "required": ["header", "items"],
        "properties": {
            "header": {
                "type": "object",
                "additionalProperties": False,
                "required": [
                    "tpa_insurer", "total_settlement_amount", "payment_mode",
                    "payment_batch", "utr_number", "settlement_number",
                    "settlement_date", "hospital_account_number",
                ],
                "properties": {
                    "tpa_insurer": {"type": ["string", "null"]},
                    "total_settlement_amount": {"type": ["number", "null"]},
                    "payment_mode": {"type": ["string", "null"]},
                    "payment_batch": {"type": ["string", "null"]},
                    "utr_number": {"type": ["string", "null"]},
                    "settlement_number": {"type": ["string", "null"]},
                    # ISO date string (YYYY-MM-DD) or null.
                    "settlement_date": {"type": ["string", "null"]},
                    "hospital_account_number": {"type": ["string", "null"]},
                },
            },
            "items": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "required": [
                        "claim_number", "settled_amount", "claim_raised_amount",
                        "disallowance", "disallowance_reason",
                    ],
                    "properties": {
                        "claim_number": {"type": ["string", "null"]},
                        "settled_amount": {"type": ["number", "null"]},
                        "claim_raised_amount": {"type": ["number", "null"]},
                        "disallowance": {"type": ["number", "null"]},
                        "disallowance_reason": {"type": ["string", "null"]},
                    },
                },
            },
        },
    },
}


_PROMPT = """You are reading an insurer/TPA settlement (remittance advice) for a hospital.

Extract TWO things and return ONE JSON object:

1. `header` — the settlement-level fields:
   - tpa_insurer: the TPA or insurer name
   - total_settlement_amount: total paid, plain number (125000, not "₹1,25,000")
   - payment_mode: e.g. NEFT, RTGS, IMPS, Cheque
   - payment_batch: any batch/lot reference for the payment
   - utr_number: the bank UTR / transaction reference
   - settlement_number: the settlement / advice number
   - settlement_date: ISO format YYYY-MM-DD
   - hospital_account_number: the hospital bank account credited
   Use null for any field not present.

2. `items` — ONE entry per claim row in the document:
   - claim_number: the claim/CCN number as written
   - settled_amount: amount settled for this claim (plain number)
   - claim_raised_amount: amount the hospital had billed/claimed (plain number)
   - disallowance: amount disallowed/deducted (plain number)
   - disallowance_reason: reason for the deduction, if stated
   Use null for any missing cell. Do NOT invent rows; only include claims
   actually present in the document.

Document content:
{text}
"""

_MAX_CHARS = 16000


def _to_text(file_bytes: bytes, file_name: str | None, content_type: str | None) -> str:
    name = (file_name or "").lower()
    ctype = (content_type or "").lower()
    try:
        if ctype == "application/pdf" or name.endswith(".pdf"):
            reader = PdfReader(io.BytesIO(file_bytes))
            return "\n".join(p.extract_text() or "" for p in reader.pages).strip()[:_MAX_CHARS]

        if name.endswith((".xlsx", ".xlsm")) or "spreadsheetml" in ctype:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            lines = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    cells = ["" if c is None else str(c) for c in row]
                    if any(cells):
                        lines.append(",".join(cells))
            return "\n".join(lines).strip()[:_MAX_CHARS]

        # CSV or any other text-ish file: decode and normalise rows.
        decoded = file_bytes.decode("utf-8", errors="replace")
        if name.endswith(".csv") or "csv" in ctype:
            reader = csv.reader(io.StringIO(decoded))
            return "\n".join(",".join(r) for r in reader).strip()[:_MAX_CHARS]
        return decoded.strip()[:_MAX_CHARS]
    except Exception as e:
        logger.error(f"Settlement file parse failed: {e}")
        return ""


def _empty() -> dict:
    return {
        "header": {
            "tpa_insurer": None, "total_settlement_amount": None, "payment_mode": None,
            "payment_batch": None, "utr_number": None, "settlement_number": None,
            "settlement_date": None, "hospital_account_number": None,
        },
        "items": [],
    }


def extract_settlement_data(file_bytes: bytes, file_name: str | None,
                            content_type: str | None) -> dict:
    """Return {header, items}. Falls back to an empty editable shape on failure."""
    text = _to_text(file_bytes, file_name, content_type)
    if not text:
        return _empty()
    if not settings.OPENAI_API_KEY:
        logger.warning("OPENAI_API_KEY not configured; skipping settlement extraction")
        return _empty()

    try:
        client = OpenAI(api_key=settings.OPENAI_API_KEY)
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": _PROMPT.format(text=text)}],
            response_format={"type": "json_schema", "json_schema": _SCHEMA},
            temperature=0,
        )
        data = json.loads(response.choices[0].message.content or "{}")
        header = data.get("header") or {}
        items = data.get("items") or []
        if not isinstance(items, list):
            items = []
        # Keep only rows that carry at least a claim number or an amount.
        cleaned = [
            {
                "claim_number": (str(it.get("claim_number")).strip() if it.get("claim_number") else None),
                "settled_amount": it.get("settled_amount"),
                "claim_raised_amount": it.get("claim_raised_amount"),
                "disallowance": it.get("disallowance"),
                "disallowance_reason": it.get("disallowance_reason"),
            }
            for it in items
            if isinstance(it, dict) and (it.get("claim_number") or it.get("settled_amount") is not None)
        ]
        return {"header": {**_empty()["header"], **header}, "items": cleaned}
    except Exception as e:
        logger.error(f"Settlement extraction failed: {e}")
        return _empty()
